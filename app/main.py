import datetime
import json
import os.path
import uuid
from enum import Enum
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from pydantic import BaseModel
from fastapi import FastAPI, HTTPException
import pandas as pd
from fastapi.responses import FileResponse, JSONResponse

app = FastAPI()


class FormaterEnum(str, Enum):
    json2xlsx = 'json2xlsx'
    dict2xlsx = 'dict2xlsx'


class TableData(BaseModel):
    rows: list[list[str]]
    columns: list[str]


class ContentRequest(BaseModel):
    title: str = None
    formater: FormaterEnum
    content: str | TableData


FORMATER_FMT = {
    FormaterEnum.json2xlsx: 'xlsx',
    FormaterEnum.dict2xlsx: 'xlsx'
}

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')


def _table_data_to_excel(table_data: dict, filepath: str):
    df = pd.DataFrame(table_data['rows'], columns=table_data['columns'])
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='测试用例')
        worksheet = writer.sheets['测试用例']
        # 行高
        for row_idx in range(1, len(df) + 2):
            worksheet.row_dimensions[row_idx].height = 48
        # 列宽
        for i in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(i)
            worksheet.column_dimensions[col_letter].width = 48

        # 自动换行
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)


@app.post('/generate-file')
async def generate_file(request: ContentRequest):
    file_id = request.title if request.title else str(uuid.uuid4())
    file_id = file_id.strip().replace(' ', '_')
    now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f'{file_id}_{now}.{FORMATER_FMT[request.formater]}'
    os.makedirs(BASE_DIR, exist_ok=True)
    filepath = os.path.join(BASE_DIR, filename)
    match request.formater:
        case FormaterEnum.json2xlsx:
            if not isinstance(request.content, str):
                raise HTTPException(400, f'无效content：{request.content}')
            try:
                data = json.loads(request.content)
            except json.decoder.JSONDecodeError as e:
                raise HTTPException(400, f'json解析失败：{e}')
            _table_data_to_excel(data, filepath)
        case FormaterEnum.dict2xlsx:
            if not isinstance(request.content, TableData):
                raise HTTPException(400, f'无效content：{request.content}')
            _table_data_to_excel(request.content.model_dump(), filepath)
        case _:
            raise HTTPException(status_code=400, detail=f'暂不支持的转换方法：{request.formater}')
    return JSONResponse({'filename': filename})


@app.get('/download-file/{filename}')
async def download_make_file(filename: str):
    filepath = os.path.join(BASE_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail=f'文件不存在或已删除：{filename}')
    return FileResponse(filepath, headers={
        'Access-Control-Allow-Origin': '*'
    })
